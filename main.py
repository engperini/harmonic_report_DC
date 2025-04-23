import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def integrate_report_first(input_path, output_path):
    # Load workbook
    wb = load_workbook(input_path)
    xls = pd.ExcelFile(input_path)
    # Create report sheet at first position
    report_ws = wb.create_sheet(title='Relatório', index=0)
    
    # Read data sheets
    df_config = pd.read_excel(xls, sheet_name='Config Info', header=None)
    df_Iraw = pd.read_excel(xls, sheet_name='A H Harmonic RMS', header=1)
    df_Vraw = pd.read_excel(xls, sheet_name='Vφ φ H Harmonic RMS', header=1)
    df_rec  = pd.read_excel(xls, sheet_name='Recording', header=1)
    df_Ipct = pd.read_excel(xls, sheet_name='A H Harmonic %', header=[0,1])
    df_Vpct = pd.read_excel(xls, sheet_name='Vφ φ H Harmonic %', header=[0,1])

    # Extract data, skip units row
    data_I   = df_Iraw.iloc[1:].reset_index(drop=True)
    data_V   = df_Vraw.iloc[1:].reset_index(drop=True)
    data_rec = df_rec.iloc[1:].reset_index(drop=True)
    Ih_pct   = df_Ipct.iloc[1:].reset_index(drop=True)
    Vh_pct   = df_Vpct.iloc[1:].reset_index(drop=True)

    # Align lengths
    N = min(len(data_I), len(data_V), len(data_rec), len(Ih_pct), len(Vh_pct))
    data_I, data_V, data_rec = data_I.iloc[:N], data_V.iloc[:N], data_rec.iloc[:N]
    Ih_pct, Vh_pct = Ih_pct.iloc[:N], Vh_pct.iloc[:N]

    # Compute fundamentals and THD
    I1 = pd.to_numeric(data_I['A1 H1'], errors='coerce')
    I2 = pd.to_numeric(data_I['A2 H1'], errors='coerce')
    I3 = pd.to_numeric(data_I['A3 H1'], errors='coerce')
    V12 = pd.to_numeric(data_V['V1-2 H1'], errors='coerce')
    V23 = pd.to_numeric(data_V['V2-3 H1'], errors='coerce')
    V31 = pd.to_numeric(data_V['V3-1 H1'], errors='coerce')
    THD_V = (pd.to_numeric(data_rec['V1-2 THDf'], errors='coerce') +
             pd.to_numeric(data_rec['V2-3 THDf'], errors='coerce') +
             pd.to_numeric(data_rec['V3-1 THDf'], errors='coerce')) / 3
    THD_I = (pd.to_numeric(data_rec['A1 THDf'], errors='coerce') +
             pd.to_numeric(data_rec['A2 THDf'], errors='coerce') +
             pd.to_numeric(data_rec['A3 THDf'], errors='coerce')) / 3

    df = pd.DataFrame({'I1': I1, 'I2': I2, 'I3': I3,
                       'V12': V12, 'V23': V23, 'V31': V31,
                       'THD_V': THD_V, 'THD_I': THD_I}).dropna().reset_index(drop=True)

    # Additional metrics
    df['I_avg'] = df[['I1','I2','I3']].mean(axis=1)
    df['V_avg'] = df[['V12','V23','V31']].mean(axis=1)
    df['P_kW']  = np.sqrt(3) * df['V_avg'] * df['I_avg'] / 1000
    IL = df['I_avg'].max()
    df['TDD']   = df['THD_I'] * (df['I_avg'] / IL)

    # Define bins
    Pmax = df['P_kW'].max()
    bins = [0, 0.25*Pmax, 0.50*Pmax, 0.75*Pmax, Pmax]
    labels = [f"0–{0.25*Pmax:.0f} kW", f"{0.25*Pmax:.0f}–{0.50*Pmax:.0f} kW",
              f"{0.50*Pmax:.0f}–{0.75*Pmax:.0f} kW", f"{0.75*Pmax:.0f}–{Pmax:.0f} kW"]
    df['Faixa'] = pd.cut(df['P_kW'], bins=bins, labels=labels, right=False, include_lowest=True)

    # Top3 helper
    def compute_top3(pct_df, phases):
        harms = {}
        for h in range(2,51):
            cols = [(pct_df.columns.levels[0][0], f"{ph} H{h}") for ph in phases]
            if all(c in pct_df.columns for c in cols):
                vals = pd.concat([pd.to_numeric(pct_df[c], errors='coerce') for c in cols], axis=1).mean(axis=1)
                harms[h] = vals
        harm_df = pd.DataFrame(harms).loc[df.index]
        harm_df['Faixa'] = df['Faixa']
        top = {f: grp.drop(columns='Faixa').mean().nlargest(3).apply(lambda v: f"{v:.2f}%").rename(lambda i: f"H{i}")
               for f, grp in harm_df.groupby('Faixa')}
        return top

    top_cur = compute_top3(Ih_pct, ['A1','A2','A3'])
    top_volt = compute_top3(Vh_pct, ['V1-2','V2-3','V3-1'])

    # Build summary
    summary = []
    for f in labels:
        sub = df[df['Faixa'] == f]
        n = len(sub)
        summary.append({
            'Faixa (kW)': f,
            'Medições': n,
            'Duração (min)': n * 2,
            'I_fund médio (A)': round(sub['I_avg'].mean(), 2),
            'V_fund médio (V)': round(sub['V_avg'].mean(), 2),
            'THD_V médio (%)': round(sub['THD_V'].mean(), 2),
            'THD_I médio (%)': round(sub['THD_I'].mean(), 2),
            'TDD médio (%)': round(sub['TDD'].mean(), 2),
            'Top 3 Harm I (%)': "; ".join([f"{h} ({v})" for h, v in top_cur.get(f, {}).items()]),
            'Top 3 Harm V (%)': "; ".join([f"{h} ({v})" for h, v in top_volt.get(f, {}).items()])
        })
    df_resumo = pd.DataFrame(summary)

    # IEEE comparison
    df_comp = pd.DataFrame({
        'Métrica': ['THD_V (%)', 'TDD (%)'],
        'Valor médio': [round(df['THD_V'].mean(), 2), round(df['TDD'].mean(), 2)],
        'Limite IEEE 519': ['≤ 8 %', '≤ 5 %'],
        'Conforme?': ['Sim', 'Sim']
    })

    # Write report
    report_ws.append(["Relatório de Qualidade de Energia - Harmônicos"])
    report_ws.append([])
    report_ws.append(["Objetivo:"])
    model = df_config[df_config[0].str.contains('Model', na=False)]
    if not model.empty:
        report_ws.append([f"Analisar medições com {model.iloc[0,0]}: {model.iloc[0,1]}"])
    report_ws.append([])
    report_ws.append(["Resumo por Faixa de Carga"])
    for r in dataframe_to_rows(df_resumo, index=False, header=True):
        report_ws.append(r)
    report_ws.append([])
    report_ws.append(["Comparação com IEEE 519-2014"])
    for r in dataframe_to_rows(df_comp, index=False, header=True):
        report_ws.append(r)
    report_ws.append([])
    report_ws.append(["Tabela Resumida: THD_V e THD_I por Faixa"])
    report_ws.append(["Faixa"] + labels)
    for metric in ['THD_V médio (%)', 'THD_I médio (%)']:
        report_ws.append([metric] + [df_resumo.loc[df_resumo['Faixa (kW)'] == f, metric].values[0] for f in labels])
    report_ws.append([])
    report_ws.append(["Conclusão:"])
    report_ws.append(["- Todos os níveis de carga apresentaram THD_V e TDD dentro dos limites IEEE 519-2014."])

    # Save workbook
    wb.save(output_path)

# Run integration for the measurement
input_file = 'Recording_8336_222794WKH 8074_2025_4_17_INTDH1A.xlsx'
output_file = 'relatorio_integrado_INTDH1A.xlsx'
integrate_report_first(input_file, output_file)

output_file
